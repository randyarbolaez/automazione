from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook, load_workbook
import os
import sys

os.chdir(os.path.expanduser("~/Desktop"))


class JobAutomation:
    """Job Class"""

    if os.path.isfile('job_progress.xlsx'):
        wb = load_workbook(filename='job_progress.xlsx')
    else:
        wb = Workbook()
    sheet = wb.active
    emptyRow = 0
    backup = 0
    jobTitle = sys.argv[1]
    locationToSearch = sys.argv[2]
    link = 'https://www.google.com/search?biw=1920&bih=1001&ei=fSfFXrKkD-rp_Qa4_53gCQ&q=site%3Agreenhouse.io+OR+site%3Ajobs.lever.co+OR+site%3Aworkable.com+OR+site%3Ajobvite.com+OR+site%3A.recruitee.com+OR+site%3Ajobs.jobvite.com%2F+OR+site%3Aboards.greenhouse.io%2F*+OR+site%3A.workable.com%2Fjobs%2F+%28' + jobTitle + '%29+AND+%28' + locationToSearch + \
        '%29&oq=site%3Agreenhouse.io+OR+site%3Ajobs.lever.co+OR+site%3Aworkable.com+OR+site%3Ajobvite.com+OR+site%3A.recruitee.com+OR+site%3Ajobs.jobvite.com%2F+OR+site%3Aboards.greenhouse.io%2F*+OR+site%3A.workable.com%2Fjobs%2F+%28' + jobTitle + '%29+AND+%28' + \
        locationToSearch + \
        '%29&gs_lcp=CgZwc3ktYWIQA1CugAFYs5IBYM2TAWgAcAB4AIABAIgBAJIBAJgBAqABAaoBB2d3cy13aXo&sclient=psy-ab&ved=0ahUKEwjy1ruuvcLpAhXqdN8KHbh_B5wQ4dUDCAw&uact=5V&num=' + \
        str(100)
    driver = webdriver.Chrome(
        ChromeDriverManager().install())
    driver.get('http://google.com/')
    driver.get(link)

    def __init__(self):
        self.listOfCompanies = []
        self.listOfJobUrl = []

    def findingEmptyRow(self):
        for i in range(1, self.sheet.max_row+2):
            cell = self.sheet.cell(row=i, column=1)
            if cell.value == None:
                self.emptyRow = i
                self.backup = i
                break
            else:
                continue

    def addJobUrlToList(self):
        for i in range(2, self.emptyRow):
            cell = self.sheet.cell(row=i, column=3)
            self.listOfJobUrl.append(cell.value)
        return self.listOfJobUrl

    def addJobToList(self):
        i = 0
        while len(self.listOfCompanies) != 5:
            try:
                if i == 101:
                    self.driver.find_element_by_xpath(
                        '//*[@id="pnnext"]').click()
                    i = 0
                else:
                    jobLink = self.driver.find_element_by_xpath(
                        '//*[@id="rso"]/div[' + str(i) + ']/div/div[1]/a').get_attribute('href')
                    name = jobLink.split('/')[3]
                if jobLink in self.listOfJobUrl:
                    print('it\'s already in excel sheet, no need to add duplicates')
                if jobLink not in self.listOfJobUrl:
                    self.listOfCompanies.append(
                        {'name': name, 'link': jobLink})
                    self.sheet['A' + str(self.emptyRow)] = name
                    self.sheet['C' + str(self.emptyRow)].hyperlink = jobLink
                    self.sheet['D' + str(self.emptyRow)] = self.jobTitle
                    self.sheet['E' + str(self.emptyRow)
                               ] = self.locationToSearch
                    self.emptyRow += 1
            except NoSuchElementException:
                print('NOT A LINK')
            i += 1

    def addCompanyUrl(self):
        print('addCompanyUrl', self.listOfCompanies)
        for j in range(len(self.listOfCompanies)):
            self.driver.get('http://google.com/')
            search = self.driver.find_element_by_xpath(
                '//*[@id="tsf"]/div[2]/div[1]/div[1]/div/div[2]/input')
            search.send_keys(self.listOfCompanies[j]['name'])
            search.submit()
            try:
                searchLink = self.driver.find_element_by_xpath(
                    '//*[@id="rso"]/div[1]/div/div/div/div[1]/a').get_attribute('href')
            except NoSuchElementException:
                searchLink = 'N/A'

            self.listOfCompanies[j]["companyUrl"] = searchLink
            self.sheet['B' + str(self.backup)].hyperlink = searchLink
            self.backup += 1

    def wholeProcess(self):
        self.findingEmptyRow()
        self.addJobUrlToList()
        self.addJobToList()
        self.addCompanyUrl()

        self.driver.quit()
        self.wb.save(filename="job_progress.xlsx")
        self.wb.close()


def main():
    j = JobAutomation()
    j.wholeProcess()


main()
